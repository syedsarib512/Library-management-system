import os
import openpyxl
from openpyxl import load_workbook
file_path = "Library-management-system\library.xlsx"
try:
    wb = load_workbook(filename=file_path)
except FileNotFoundError:
    wb = openpyxl.Workbook()
    sheet = wb.active
    admin = wb.create_sheet("Admin Data")
    books = wb.create_sheet("Books")
    books['A1'] = "ID"
    books['B1'] = "Tile"
    books['C1'] = 'Author'
    books['D1'] = "ISBN"
    books['E1'] = "Available Copies"
    books['F1'] = 'Issued By'
    admin['A1'] = "User Id"
    admin['B1'] = "User Name"
    admin['C1'] = "Password"
    sheet['A1'] = "User Id"
    sheet['B1'] = "User Name"
    sheet['C1'] = "Password"
    sheet["D1"] = "Status"
    wb.save(file_path)
class User:
    def __init__(self, user_id, username, password):
        for i in range(1,100):
            if wb['Sheet'].cell(row=i,column=1).value == user_id:
                break
            if wb['Sheet'].cell(row=i,column=1).value == None:
                wb['Sheet'].cell(row=i,column=1).value = user_id
                wb['Sheet'].cell(row=i,column=2).value = username
                wb['Sheet'].cell(row=i,column=3).value = password
                wb.save(file_path)
                break
    def LogIn(self,user_id,username,password):
        for i in range(1,100):
            if wb['Sheet'].cell(row=i,column=1).value == user_id:
                if wb['Sheet'].cell(row=i,column=2).value == username:
                    if wb['Sheet'].cell(row=i,column=3).value == password:
                        print("Log In as Student Successfully")
                        self.login_id_user = user_id
                        self.Menu()
                    else:
                        print('Wrong Password!')
                else:
                    print('Invalid Username!')
        print("Invalid Id!")
    def Issue_Book(self,book_id,ISBN):
        found = False
        for i in range(1,100):
            if wb['Sheet'].cell(row=i,column=1).value == self.login_id_user:
                if wb['Sheet'].cell(row=i,column=4).value != None:
                    print("You are not allowed to Issue Book")
                    self.Menu()
        for i in range(1,100):
            if wb['Books'].cell(row=i,column=1).value == book_id:
                if wb['Books'].cell(row=i,column=4).value == ISBN:
                    if wb['Books'].cell(row=i,column=5).value == 1:
                        wb['Books'].cell(row=i,column=5).value == 0
                        wb["Books"].cell(row=i,column=6).value = self.login_id_user
                        wb.save(file_path)
                        print("Book Issued Successfully!")
                        Found = True
                        break
                    else:
                        print("Book is not Available")
                else:
                    print("Invalid ISBN")
        if found == False:
            print("No Id Matched")
    def Return_Book(self,book_id):
        for i in range(1,100):
            if wb['Books'].cell(row=i,column=1).value == book_id:
                if wb['Books'].cell(row=i,column=6).value == self.login_id_user:
                    wb['Books'].cell(row=i,column=6).value = None
                    wb['Books'].cell(row=i,column=5).value = 1
                    print("Book Returned Successfully!")
                else:
                    print("You didn't issued this book")
            print("No Book Found with this id")
    def Menu(self):
        print("Welcome to User Menu\nWhat do you want\n1 Issue Book \n2 Return Book \n3 Back to HomePage")
        choice = input("Enter your choice here : ")
        if choice == "1":
            book_id = input("Enter the Book Id : ")
            ISBN = input("Enter the ISBN : ")
            self.Issue_Book(book_id,ISBN)
            self.Menu()
        elif choice == '2':
            book_id = input("Enter the Book Id : ")
            self.Return_Book(book_id)
            self.Menu()
        elif choice == '3':
            pass
        else:
            print("Invalid Input! Try Again!")
            self.Menu()
class Librarian:
    def __init__(self, user_id, username, password):
        for i in range(1,100):
            if wb['Admin Data'].cell(row=i,column=1).value == user_id:
                break
            if wb['Admin Data'].cell(row=i,column=1).value == None:
                wb['Admin Data'].cell(row=i,column=1).value = user_id
                wb['Admin Data'].cell(row=i,column=2).value = username
                wb['Admin Data'].cell(row=i,column=3).value = password
                wb.save(file_path)
                break
    def LogIn(self,user_id,username,password):
        if wb['Admin Data'].cell(row=2,column=1).value == user_id:
            if wb['Admin Data'].cell(row=2,column=2).value == username:
                if wb['Admin Data'].cell(row=2,column=3).value == password:
                    print("Log In as User Successfully")
                    self.Menu()
                else:
                    print('Wrong Password!')
            else:
                print('Invalid Username!')
        else:
            print("Invalid Id!")
    def Block_User(self,user_id):
        for i in range(1,100):
            if wb['Sheet'].cell(row=i,column=1).value == user_id:
                wb['Sheet'].cell(row=i,column=4).value = "Blocked"
                wb.save(file_path)
                print("User Blocked Successfully!")
    def Unblock_User(self,user_id):
        for i in range(1,100):
            if wb['Sheet'].cell(row=i,column=1).value == user_id:
                wb['Sheet'].cell(row=i,column=4).value = None
                wb.save(file_path)
                print("User UnBlocked Successfully!")
    def Menu(self):
        print("Welcome to Admin Menu")
        print("What do you want \n1 Add User \n2 Block User \n3 Unblock User \n4 Back to Homepage")
        choice = input("Enter your Choice Here : ")
        if choice == '1':
            user_id = input("Enter User Id : ")
            user_name = input("Enter User name : ")
            password = input("Enter Password : ")
            User(user_id,user_name,password)
            self.Menu()
        elif choice == '2':
            user_id = input("Enter User Id : ")
            self.Block_User(user_id)
            self.Menu()
        elif choice == '3':
            user_id = input("Enter User Id : ")
            self.Unblock_User(user_id)
            self.Menu()
        elif choice == '4':
            pass
        else:
            print("Invalid Input! Try Again!")
            self.Menu()
class Book:
    def __init__(self, book_id, title, author, ISBN, available_copies):
        for i in range(1,1000):
            if wb['Books'].cell(row=i,column=1).value == book_id:
                break
            if wb['Books'].cell(row=i,column=1).value == None:
                wb['Books'].cell(row=i,column=1).value = book_id
                wb['Books'].cell(row=i,column=2).value = title
                wb['Books'].cell(row=i,column=3).value = author
                wb['Books'].cell(row=i,column=4).value = ISBN
                wb['Books'].cell(row=i,column=5).value = available_copies
                wb.save(file_path)
                print("Book Added Successfully!")
                break
class ConsoleUI:
    def __init__(self, library,user):
        self.library = library
        self.user = user
    def run(self):
        print("Wellcome To Console Based Library Management Sysytem")
        print("1. Librarian\n 2. Student or Teacher")
        start = input("Enter your Choice here : ")
        if start == "1":
            user_id = input("Enter user id : ")
            username = input("Enter username here : ")
            password = input("Enter Password : ")
            Librarian.LogIn(self.library,user_id,username,password)
            self.run()
        elif start == '2':
            user_id = input("Enter user id : ")
            username = input("Enter username here : ")
            password = input("Enter Password : ")
            User.LogIn(self.user,user_id,username,password)
            self.run()
        else:
            print("Invalid Input. Try Again!")
            self.run()
librarian = Librarian("1", "admin", "admin123")
user1 = User("2", "user1", "user123")

# Adding Books
famous_books = [
    {"title": "To Kill a Mockingbird", "author": "Harper Lee", "ISBN": "978-0-06-112008-4"},
    {"title": "1984", "author": "George Orwell", "ISBN": "978-0-452-28423-4"},
    {"title": "Pride and Prejudice", "author": "Jane Austen", "ISBN": "978-0-486-45468-5"},
    {"title": "The Great Gatsby", "author": "F. Scott Fitzgerald", "ISBN": "978-0-7432-7356-5"},
    {"title": "Moby-Dick", "author": "Herman Melville", "ISBN": "978-1-4209-3207-1"},
    {"title": "War and Peace", "author": "Leo Tolstoy", "ISBN": "978-0-14-044793-4"},
    {"title": "The Catcher in the Rye", "author": "J.D. Salinger", "ISBN": "978-0-316-76948-7"},
    {"title": "To the Lighthouse", "author": "Virginia Woolf", "ISBN": "978-0-15-690739-2"},
    {"title": "The Odyssey", "author": "Homer", "ISBN": "978-0-19-923332-6"},
    {"title": "One Hundred Years of Solitude", "author": "Gabriel García Márquez", "ISBN": "978-0-06-088328-7"},
    {"title": "Crime and Punishment", "author": "Fyodor Dostoevsky", "ISBN": "978-0-14-044913-6"},
    {"title": "The Lord of the Rings", "author": "J.R.R. Tolkien", "ISBN": "978-0-544-27334-7"},
    {"title": "Jane Eyre", "author": "Charlotte Brontë", "ISBN": "978-0-553-21328-1"},
    {"title": "The Hobbit", "author": "J.R.R. Tolkien", "ISBN": "978-0-618-00221-9"},
    {"title": "The Chronicles of Narnia", "author": "C.S. Lewis", "ISBN": "978-0-06-623850-0"},
    {"title": "Brave New World", "author": "Aldous Huxley", "ISBN": "978-0-06-085052-4"},
    {"title": "The Grapes of Wrath", "author": "John Steinbeck", "ISBN": "978-0-14-200066-3"},
    {"title": "The Adventures of Huckleberry Finn", "author": "Mark Twain", "ISBN": "978-0-486-44085-4"},
    {"title": "Fahrenheit 451", "author": "Ray Bradbury", "ISBN": "978-1-4516-7331-9"},
    {"title": "The Alchemist", "author": "Paulo Coelho", "ISBN": "978-0-06-112241-5"},
    {"title": "The Brothers Karamazov", "author": "Fyodor Dostoevsky", "ISBN": "978-0-14-044109-3"},
    {"title": "Wuthering Heights", "author": "Emily Brontë", "ISBN": "978-0-553-21305-2"},
    {"title": "Don Quixote", "author": "Miguel de Cervantes", "ISBN": "978-0-14-243723-0"},
    {"title": "The Picture of Dorian Gray", "author": "Oscar Wilde", "ISBN": "978-0-486-42129-7"},
    {"title": "Dracula", "author": "Bram Stoker", "ISBN": "978-0-553-21311-3"},
    {"title": "The Sun Also Rises", "author": "Ernest Hemingway", "ISBN": "978-0-684-80122-3"},
    {"title": "Anna Karenina", "author": "Leo Tolstoy", "ISBN": "978-0-14-303500-8"},
    {"title": "The Old Man and the Sea", "author": "Ernest Hemingway", "ISBN": "978-0-684-80122-3"},
    {"title": "A Tale of Two Cities", "author": "Charles Dickens", "ISBN": "978-0-14-143960-0"},
    {"title": "Moby-Dick", "author": "Herman Melville", "ISBN": "978-1-4209-3207-1"},
    {"title": "War and Peace", "author": "Leo Tolstoy", "ISBN": "978-0-14-044793-4"},
    {"title": "The Catcher in the Rye", "author": "J.D. Salinger", "ISBN": "978-0-316-76948-7"},
    {"title": "To the Lighthouse", "author": "Virginia Woolf", "ISBN": "978-0-15-690739-2"},
    {"title": "The Odyssey", "author": "Homer", "ISBN": "978-0-19-923332-6"},
    {"title": "One Hundred Years of Solitude", "author": "Gabriel García Márquez", "ISBN": "978-0-06-088328-7"},
    {"title": "Crime and Punishment", "author": "Fyodor Dostoevsky", "ISBN": "978-0-14-044913-6"},
    {"title": "The Lord of the Rings", "author": "J.R.R. Tolkien", "ISBN": "978-0-544-27334-7"},
    {"title": "Jane Eyre", "author": "Charlotte Brontë", "ISBN": "978-0-553-21328-1"},
    {"title": "The Hobbit", "author": "J.R.R. Tolkien", "ISBN": "978-0-618-00221-9"},
    {"title": "The Chronicles of Narnia", "author": "C.S. Lewis", "ISBN": "978-0-06-623850-0"},
    {"title": "Brave New World", "author": "Aldous Huxley", "ISBN": "978-0-06-085052-4"},
    {"title": "The Grapes of Wrath", "author": "John Steinbeck", "ISBN": "978-0-14-200066-3"},
    {"title": "The Adventures of Huckleberry Finn", "author": "Mark Twain", "ISBN": "978-0-486-44085-4"},
    {"title": "Fahrenheit 451", "author": "Ray Bradbury", "ISBN": "978-1-4516-7331-9"},
    {"title": "The Alchemist", "author": "Paulo Coelho", "ISBN": "978-0-06-112241-5"},
    {"title": "The Brothers Karamazov", "author": "Fyodor Dostoevsky", "ISBN": "978-0-14-044109-3"},
    {"title": "Wuthering Heights", "author": "Emily Brontë", "ISBN": "978-0-553-21305-2"},
    {"title": "Don Quixote", "author": "Miguel de Cervantes", "ISBN": "978-0-14-243723-0"},
    {"title": "The Picture of Dorian Gray", "author": "Oscar Wilde", "ISBN": "978-0-486-42129-7"},
    {"title": "Dracula", "author": "Bram Stoker", "ISBN": "978-0-553-21311-3"},
    {"title": "The Sun Also Rises", "author": "Ernest Hemingway", "ISBN": "978-0-684-80122-3"},
    {"title": "Anna Karenina", "author": "Leo Tolstoy", "ISBN": "978-0-14-303500-8"},
    {"title": "The Old Man and the Sea", "author": "Ernest Hemingway", "ISBN": "978-0-684-80122-3"},
    {"title": "A Tale of Two Cities", "author": "Charles Dickens", "ISBN": "978-0-14-143960-0"},
    {"title": "Moby-Dick", "author": "Herman Melville", "ISBN": "978-1-4209-3207-1"},
    {"title": "War and Peace", "author": "Leo Tolstoy", "ISBN": "978-0-14-044793-4"},
    {"title": "The Catcher in the Rye", "author": "J.D. Salinger", "ISBN": "978-0-316-76948-7"},
    {"title": "To the Lighthouse", "author": "Virginia Woolf", "ISBN": "978-0-15-690739-2"},
    {"title": "The Odyssey", "author": "Homer", "ISBN": "978-0-19-923332-6"},
    {"title": "One Hundred Years of Solitude", "author": "Gabriel García Márquez", "ISBN": "978-0-06-088328-7"},
    {"title": "Crime and Punishment", "author": "Fyodor Dostoevsky", "ISBN": "978-0-14-044913-6"},
    {"title": "The Lord of the Rings", "author": "J.R.R. Tolkien", "ISBN": "978-0-544-27334-7"},
    {"title": "Jane Eyre", "author": "Charlotte Brontë", "ISBN": "978-0-553-21328-1"},
    {"title": "The Hobbit", "author": "J.R.R. Tolkien", "ISBN": "978-0-618-00221-9"},
    {"title": "The Chronicles of Narnia", "author": "C.S. Lewis", "ISBN": "978-0-06-623850-0"},
    {"title": "Brave New World", "author": "Aldous Huxley", "ISBN": "978-0-06-085052-4"},
    {"title": "The Grapes of Wrath", "author": "John Steinbeck", "ISBN": "978-0-14-200066-3"},
    {"title": "The Adventures of Huckleberry Finn", "author": "Mark Twain", "ISBN": "978-0-486-44085-4"},
    {"title": "Fahrenheit 451", "author": "Ray Bradbury", "ISBN": "978-1-4516-7331-9"},
    {"title": "The Alchemist", "author": "Paulo Coelho", "ISBN": "978-0-06-112241-5"},
    {"title": "The Brothers Karamazov", "author": "Fyodor Dostoevsky", "ISBN": "978-0-14-044109-3"},
    {"title": "Wuthering Heights", "author": "Emily Brontë", "ISBN": "978-0-553-21305-2"},
    {"title": "Don Quixote", "author": "Miguel de Cervantes", "ISBN": "978-0-14-243723-0"},
    {"title": "The Picture of Dorian Gray", "author": "Oscar Wilde", "ISBN": "978-0-486-42129-7"},
    {"title": "Dracula", "author": "Bram Stoker", "ISBN": "978-0-553-21311-3"},
    {"title": "The Sun Also Rises", "author": "Ernest Hemingway", "ISBN": "978-0-684-80122-3"},
    {"title": "Anna Karenina", "author": "Leo Tolstoy", "ISBN": "978-0-14-303500-8"},
    {"title": "The Old Man and the Sea", "author": "Ernest Hemingway", "ISBN": "978-0-684-80122-3"},
    {"title": "A Tale of Two Cities", "author": "Charles Dickens", "ISBN": "978-0-14-143960-0"},
]

for i, book_info in enumerate(famous_books, start=1):
    Book(
        book_id=str(i),
        title=book_info["title"],
        author=book_info["author"],
        ISBN=book_info["ISBN"],
        available_copies=1
    )

#library.add_user(librarian)
#library.add_user(user1)
#library.add_book(book1)
# Initialize the console-based UI and run the system
ui = ConsoleUI(librarian,user1)
ui.run()



